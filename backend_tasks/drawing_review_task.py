import argparse
import openpyxl
import os
import sys
from pathlib import Path
from pdf2image import convert_from_path
from vertexai.generative_models import GenerativeModel
from utils.simple_multi_genemipronpt import (
    generate_with_multiple_contents,
)
from utils.gemini_response import (
    get_raw_response,
)


# サービスアカウントキーファイルのパスを設定
os.environ['GOOGLE_APPLICATION_CREDENTIALS'] = 'poc-shared-mab-ai-adv-util-sa.json'
model = GenerativeModel("gemini-2.5-pro")


def check_drawings_prompt(propt_list, files_list):
    # 反映前の図面について確認
    bf_check_drawings_prompt = f"""
        あなたは、製図図面を読み取るためのアシスタントです。
        {files_list['before_file']}は製図図面です。
        {propt_list['bf_grid']}付近にある寸法情報について出力してください。
        ルール
            1:JSON形式でにて出力してください。
            "view_part": "正面図",//指定したgrid座標付近にある
            "dimension"://寸法や設計指示ごとに作成
                "field": "直径寸法",//どのような寸法かもしくは品質指示および注記なのか。また、付近に管理番号がある場合は"()"がついた数字で記載してください
                "value": "3× 20.3±0.08",//寸法値や設計指示。"※"がある場合は寸法のあとつける。寸法指示が別途仕様表に記載があるため、項目と同一記載の寸法を探して記載
                "notice": ”全幅”,//他の寸法との関係性がわかる形で、図面内でどのような寸法の値かを記載。また、"※"印がある場合、対応する注記の内容を記載
            2:出力はjson形式のみにしてください
    """

    raw_bf_check_drawings = generate_with_multiple_contents(
        model,
        image_paths=[f"{files_list['before_file']}"],
        prompts=[bf_check_drawings_prompt],
    )

    bf_check_drawings = get_raw_response(raw_bf_check_drawings)

    # print(f"反映前情報出力: {bf_check_drawings}")

    # 反映前の図面と指摘の情報から反映事項の予測
    predict_prompt = f"""
        あなたは、製図図面を確認するアシスタントです。
        {files_list['before_file']}の製図図面内の下記の指摘がどのように反映されるべきか教えてください。
        下記は指摘事項と指摘部分に付近の寸法情報です。

        {propt_list['check_item']}

        {propt_list['bf_grid']}付近の寸法情報
        {bf_check_drawings}
        ルール
            1:JSON形式でにて出力してください。
                "修正前の状態": ,//指摘事項と寸法情報から予測できる修正前の状態
                "修正後の状態": ,//指摘事項と寸法情報から予測できる修正後の状態
                "指摘箇所": ,//指摘事項の場所
                "修正内容": ,//どのような修正が行われたか記載
            2:出力はjson形式のみにしてください
    """
    predict_images = [f"{files_list['before_file']}"]
    refer_path = files_list.get("refer_file")
    if refer_path:
        predict_images.append(f"{refer_path}")

    raw_check_predict = generate_with_multiple_contents(
        model,
        image_paths=predict_images,
        prompts=[predict_prompt],
    )
    check_predict = get_raw_response(raw_check_predict)

    # print(f"指摘事項反映内容予測出力: {check_predict}")

    # 反映後の図面について確認
    af_check_drawings_prompt = f"""
        あなたは、製図図面を読み取るためのアシスタントです。
        {files_list['after_file']}は製図図面です。
        {propt_list['af_grid']}付近にある寸法情報について出力してください。
        ルール
            1:JSON形式でにて出力してください。
            "view_part": "正面図",//指定したgrid座標付近にある
            "dimension"://寸法や設計指示ごとに作成
                "field": "直径寸法",//どのような寸法かもしくは品質指示および注記なのか。また、付近に管理番号がある場合は"()"がついた数字で記載してください
                "value": "3× 20.3±0.08",//寸法値や設計指示。"※"がある場合は寸法のあとつける。寸法指示が別途仕様表に記載があるため、項目と同一記載の寸法を探して記載
                "notice": ”全幅”,//他の寸法との関係性がわかる形で、図面内でどのような寸法の値かを記載。また、"※"印がある場合、対応する注記の内容を記載
            2:出力はjson形式のみにしてください
    """
    raw_af_check_drawings = generate_with_multiple_contents(
        model,
        image_paths=[f"{files_list['after_file']}"],
        prompts=[af_check_drawings_prompt],
    )

    af_check_drawings = get_raw_response(raw_af_check_drawings)

    # print(f"反映後情報出力: {af_check_drawings}")

    judge_prompt = f"""
        あなたは、製図図面を確認するアシスタントです。
        下記に反映すべき事項と付近の寸法情報です。
        反映すべき事項内の”修正後の状態”が正しいと仮定して、{files_list['after_file']}図面に反映されているかどうかを、理由を踏まえて教えてください。

        反映すべき事項:
        {check_predict}

        {propt_list['af_grid']}付近の寸法情報
        {af_check_drawings}
    """
    raw_judge_check_drawings = generate_with_multiple_contents(
        model,
        image_paths=[f"{files_list['after_file']}"],
        prompts=[judge_prompt],
    )

    judge_check_drawings = get_raw_response(raw_judge_check_drawings)

    # print(f"情報判断出力: {judge_check_drawings}")

    return judge_check_drawings


def read_excel_to_list(excel_path: str) -> list[list]:
    """Excelを読み込む。必要であればデータをフィルタリング"""
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["図面審査シート"]

    target_text = "①各部門から指摘・要望事項　（各部門で記入）"
    found_cell = None

    # 文字列の検索
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == target_text:
                found_cell = cell
                break
        if found_cell:
            break

    extracted_data = []
    if found_cell:
        start_row_index = found_cell.row + 2
        # start_col_index = found_cell.column
        start_col_index = 1

        print(f"基準セル: {found_cell.coordinate}")
        print(
            f"Excelの取得開始行: {start_row_index}, Excelの取得開始列: {start_col_index}")

        # 必要なデータのみに絞り込む
        for row in ws.iter_rows(min_row=start_row_index, min_col=start_col_index, values_only=True):
            no = row[0]
            category = row[1]
            product_name = row[2]
            bf_code = row[3]
            bf_grid = row[4]
            issue_and_request = row[5]
            is_use = row[6]
            af_code = row[7]
            af_grid = row[8]
            evaluation_result = row[9]
            if is_use == "可":
                extracted_data.append([no, category, product_name, bf_code, bf_grid,
                                      issue_and_request, is_use, af_code, af_grid, evaluation_result])
    else:
        print("見つかりません")

    wb.close()

    return extracted_data


def write_result(excel_path: str, save_dir: str, data: list[list]) -> None:
    wb = openpyxl.load_workbook(excel_path)

    file_name = f"result_{Path(excel_path).stem}.xlsx"
    print(f"アウトプット用のファイル名: {file_name}")

    save_sheet_name = "AI判定結果"
    print(wb.sheetnames)
    if save_sheet_name in wb.sheetnames:
        wb.remove(wb[save_sheet_name])

    wb.create_sheet(title=save_sheet_name)
    ws = wb[save_sheet_name]
    ws["A1"] = "No"
    ws["B1"] = "審査部門"
    ws["C1"] = "品名"
    ws["D1"] = "指摘先図番 CODE."
    ws["E1"] = "指摘先座標"
    ws["F1"] = "指摘・要望事項"
    ws["G1"] = "採用可否"
    ws["H1"] = "反映先図番 CODE."
    ws["I1"] = "反映先座標"
    ws["J1"] = "関連部門との検討結果"
    ws["K1"] = "AI判定結果"
    ws["L1"] = "判定理由"
    for row_list in data:
        ws.append(row_list)

    print(save_dir / file_name)

    wb.save(save_dir / file_name)
    wb.close()
    return None


def pdf_to_jpeg(file_path):
    """PDFを画像に変換する"""
    file_path = Path(file_path)

    images = convert_from_path(file_path)

    save_path = file_path.with_suffix('.jpg')

    # 各ページを画像として保存する
    for i, image in enumerate(images):
        image.save(save_path, 'JPEG')

    return save_path


if __name__ == "__main__":
    parse = argparse.ArgumentParser("図面審査")
    parse.add_argument(
        "--excel-dir",
        type=str,
        help="Excelファイルのパス"
    )
    parse.add_argument(
        "--pdf-dir",
        type=str,
        help="PDFの図面のパス"
    )
    parse.add_argument(
        "--output-dir",
        type=str,
        help="実行結果出力先ディレクトリ"
    )
    args = parse.parse_args()

    if not os.path.isdir(args.output_dir):
        os.makedirs(args.output_dir)
        print(f"{args.output_dir}を作成しました")

    output_dir = Path(args.output_dir)

    excel_dir = Path(args.excel_dir)
    excel_path = list(excel_dir.glob("*.xlsx"))[0]
    print(f"対象のExcelファイル: {excel_path}")

    extracted_data = read_excel_to_list(excel_path=excel_path)

    pdf_dir = Path(args.pdf_dir)
    pdf_files = list(pdf_dir.glob("*.pdf"))

    image_files = [pdf_to_jpeg(file) for file in pdf_files]

    results = []
    for i in extracted_data:
        print(f"row: {i}")
        propt_list = {}
        files_list = {}

        # ファイル名検索
        p = Path(args.pdf_dir)
        bf_files = [f for f in p.glob(f"{i[0]}*bf_file_*{i[3]}*.jpg")]
        af_files = [f for f in p.glob(f"{i[0]}*af_file_*{i[7]}*.jpg")]
        print(f"bf_files: {bf_files}")
        print(f"af_files: {af_files}")
        if len(bf_files) == 1 and len(af_files) == 1:
            bf_file = bf_files[0]
            af_file = af_files[0]
        else:
            # TODO: 暫定対応
            continue

        # TODO: 暫定対応_各種値を格納
        files_list["before_file"] = bf_file
        files_list["after_file"] = af_file

        propt_list["bf_grid"] = i[4]
        propt_list["af_grid"] = i[8]
        propt_list["check_item"] = f"""
        指摘事項
            {i[5]}

        回答
            {i[9]}
        """

        res = check_drawings_prompt(propt_list, files_list)
        # print(f"レスポンス: {res}")

        # 反映状況
        check = "反映済" if "反映されています" in res else "未反映"
        i.append(check)

        # 反映理由
        keyword = "理由"
        start_index = res.find(keyword)
        extracted_res = res[start_index:]
        i.append(extracted_res)

        results.append(i)
        # print(f"結果: {results}")

    # TODO: 処理が終わった後にExcelシートを作成
    write_result(excel_path, output_dir, results)

    sys.exit(0)
